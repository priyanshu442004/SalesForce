import os
import pandas as pd
import numpy as np
import datetime
import traceback

def is_column_all_empty_or_null(series):
    """
    Checks if a pandas Series is entirely composed of empty, NaN, or string 'NULL' / 'None' values.
    """
    if series.isna().all():
        return True
    
    # Standardize as string for text null checks
    str_series = series.astype(str).str.strip().str.upper()
    empty_or_null_mask = str_series.isin(['', 'NAN', 'NULL', 'NONE', '<NA>', 'N/A']) | series.isna()
    return empty_or_null_mask.all()

def format_date_value(val, format_str):
    """
    Parses date and formats it based on YYYY-MM-DD and YYYY-MM-DDTHH:MM:SS.000Z styles.
    """
    if pd.isna(val) or str(val).strip() in ['', 'nan', 'NAN', 'None', 'NULL']:
        return None
    
    try:
        # Convert to pandas datetime
        if isinstance(val, (datetime.datetime, datetime.date)):
            dt = val
        else:
            dt = pd.to_datetime(val, errors='coerce')
            if pd.isna(dt):
                return val # fallback to original if parsing fails
            
        # Convert Excel YYYY-MM-DD format to Python's %Y-%m-%d strftime format
        py_fmt = str(format_str).replace('YYYY', '%Y').replace('DD', '%d')
        py_fmt = py_fmt.replace('HH', '%H').replace('SS', '%S')
        py_fmt = py_fmt.replace('hh', '%H').replace('ss', '%S')
        
        # Replace first 'MM' with '%m' (Month) and any remaining 'MM' with '%M' (Minute)
        if 'MM' in py_fmt:
            py_fmt = py_fmt.replace('MM', '%m', 1).replace('MM', '%M')
            
        # Support lowercase 'mm' for minutes
        py_fmt = py_fmt.replace('mm', '%M')
        
        # Manually format ISO millisecond suffix if specified literally
        if '.000Z' in py_fmt:
            base_fmt = py_fmt.replace('.000Z', '')
            return dt.strftime(base_fmt) + '.000Z'
            
        return dt.strftime(py_fmt)
    except Exception as e:
        print(f"Error formatting date {val} with format {format_str}: {e}")
        return val

def process_preview(source_path: str, master_path: str, logic_path: str, output_path: str) -> dict:
    """
    Loads migration Excel sheets, cleans empty columns, resolves Salesforce master lookups, 
    formats dates in-place, and outputs preview.xlsx. Handles any columns, files, and schemas dynamically.
    Supports fallback substring matching for robust joins (e.g. matching parts of names/ids).
    Formats the downloaded output sheet to an executive, premium enterprise style.
    """
    try:
        # 1. Verify existence of files
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"Source file not found at {source_path}")
        if not os.path.exists(master_path):
            raise FileNotFoundError(f"Salesforce Master file not found at {master_path}")
        if not os.path.exists(logic_path):
            raise FileNotFoundError(f"Mapping Logic file not found at {logic_path}")
            
        print("Loading Source Data and Mapping Logic sheets...")
        source_df = pd.read_excel(source_path)
        
        # Read Mapping Logic
        logic_excel = pd.ExcelFile(logic_path)
        # Always identify and read the very first sheet inside the mapping excel
        wo_sheet = logic_excel.sheet_names[0]
        print(f"Reading first sheet for mapping logic: '{wo_sheet}'")
            
        logic_df = pd.read_excel(logic_path, sheet_name=wo_sheet)
        
        # 2. Dynamic Column Mappings Resolver for Mapping Logic File
        # Resolves any variations (e.g. trailing/leading spaces or separate master sheet and format columns)
        col_mapping = {}
        for c in logic_df.columns:
            c_clean = str(c).strip().lower()
            if c_clean in ["source fields", "source field"]:
                col_mapping["source_field"] = c
            elif c_clean in ["target fields", "target field", "netsuits"]:
                col_mapping["target_field"] = c
            elif c_clean == "data type":
                col_mapping["data_type"] = c
            elif c_clean == "default + format" or c_clean == "format":
                col_mapping["rule"] = c
            elif c_clean == "master sheet" or c_clean == "rule" or c_clean == "master file":
                col_mapping["master_sheet"] = c
            elif c_clean == "search column" or c_clean == "pk":
                col_mapping["search_col"] = c
            elif c_clean == "copyable column" or c_clean == "fk":
                col_mapping["copyable_col"] = c
            elif c_clean in ["comment", "comments"]:
                col_mapping["comments"] = c

        # Fallbacks to loose matching if any key column was not resolved
        for c in logic_df.columns:
            c_clean = str(c).strip().lower()
            if "source_field" not in col_mapping and "source field" in c_clean:
                col_mapping["source_field"] = c
            if "target_field" not in col_mapping and ("target field" in c_clean or "netsuits" in c_clean):
                col_mapping["target_field"] = c
            if "data_type" not in col_mapping and "data type" in c_clean:
                col_mapping["data_type"] = c
            if "rule" not in col_mapping and ("default + format" in c_clean or "format" in c_clean):
                col_mapping["rule"] = c
            if "master_sheet" not in col_mapping and ("master sheet" in c_clean or "rule" in c_clean or "master file" in c_clean):
                col_mapping["master_sheet"] = c
            if "search_col" not in col_mapping and ("search column" in c_clean or "pk" in c_clean):
                col_mapping["search_col"] = c
            if "copyable_col" not in col_mapping and ("copyable column" in c_clean or "fk" in c_clean):
                col_mapping["copyable_col"] = c
            if "comments" not in col_mapping and "comment" in c_clean:
                col_mapping["comments"] = c

        print("Resolved Mapping Logic columns:")
        for k, v in col_mapping.items():
            print(f"  - {k} -> '{v}'")

        # Fallbacks for missing column definitions
        get_val = lambda row, key, default="": str(row.get(col_mapping.get(key, ""), default)).strip()

        # Extract values supporting both split columns and single combined column layouts
        def get_rule_and_sheet_vals(row):
            master_sheet_col = col_mapping.get("master_sheet")
            rule_col = col_mapping.get("rule")
            
            master_sheet_val = str(row.get(master_sheet_col, "")).strip() if master_sheet_col else ""
            rule_val = str(row.get(rule_col, "")).strip() if rule_col else master_sheet_val
            
            # If they were mapped to the same single column, standardise
            if master_sheet_col == rule_col:
                rule_val = master_sheet_val
                
            return master_sheet_val, rule_val

        # 3. Identify and Load ONLY required Master Sheets
        # This keeps the parsing lightning-fast regardless of the master file size
        required_sheets = set()
        for _, rule in logic_df.iterrows():
            data_type_val = get_val(rule, "data_type").lower()
            master_sheet_val, rule_val = get_rule_and_sheet_vals(rule)
            
            # Sheet name resides in master_sheet_val or rule_val depending on how it's formatted
            sheet_ref = master_sheet_val if master_sheet_val and master_sheet_val.lower() != "nan" else rule_val
            
            if data_type_val == "lookup" and sheet_ref and sheet_ref.lower() != "nan":
                if ">" in sheet_ref:
                    sheet_name = sheet_ref.split(">")[0].strip()
                else:
                    sheet_name = sheet_ref
                required_sheets.add(sheet_name.lower().strip())
                
        print(f"Optimized Master sheets to load: {required_sheets}")
        
        master_sheets = {}
        master_excel = pd.ExcelFile(master_path)
        for s_name in master_excel.sheet_names:
            s_name_clean = s_name.strip().lower()
            if s_name_clean in required_sheets:
                print(f"Loading required master sheet: '{s_name}'")
                master_sheets[s_name_clean] = pd.read_excel(master_path, sheet_name=s_name)

        # 4. Keep all columns from source file (no filtering of columns whose values are all blank or null)
        columns_list = []
        
        # Add "uniqueID" as the first column
        num_rows = len(source_df)
        unique_ids = list(range(1, num_rows + 1))
        columns_list.append({
            "name": "uniqueID",
            "values": unique_ids,
            "type": "original",
            "meta": "ID"
        })
        
        for col in source_df.columns:
            vals = source_df[col].replace({np.nan: None}).tolist()
            columns_list.append({
                "name": str(col).strip(),
                "values": vals,
                "type": "original"
            })
            
        dropped_columns = []
        print("Retained all source columns (no filtering of empty/NULL columns). Added 'uniqueID' as the first column.")
            
        # Stats summary counters
        lookups_attempted = 0
        lookups_successful = 0
        dates_formatted = 0
        new_fields_added = 0
        
        # Helper to get column values from columns_list by name
        def get_original_col_values(col_name):
            col_name_stripped = col_name.strip().lower()
            for col in columns_list:
                if col["name"].strip().lower() == col_name_stripped and col["type"] == "original":
                    return col["values"]
            return None
            
        # Robust schema-agnostic helper to locate any ID-like columns in source data
        def get_carelink_id_values():
            # 1. Search explicitly for Carelink ID
            for col in columns_list:
                n = col["name"].strip().lower()
                if "carelink" in n and "id" in n:
                    return col["values"]
            # 2. Search for any Carelink
            for col in columns_list:
                n = col["name"].strip().lower()
                if "carelink" in n:
                    return col["values"]
            # 3. Search for any ID-like columns
            for col in columns_list:
                n = col["name"].strip().lower()
                if "id" in n:
                    return col["values"]
            # 4. Fallback to the first column of original data
            for col in columns_list:
                if col["type"] == "original":
                    return col["values"]
            return None

        # 5. Iterate and Apply Mapping Rules
        print("Applying mapping and transformation rules...")
        for _, rule in logic_df.iterrows():
            source_field = get_val(rule, "source_field")
            target_field = get_val(rule, "target_field")
            data_type = get_val(rule, "data_type").lower()
            search_col_name = get_val(rule, "search_col")
            copyable_col_name = get_val(rule, "copyable_col")
            
            master_sheet_val, rule_val = get_rule_and_sheet_vals(rule)
            
            # Skip empty or spacer rows
            if not source_field or source_field.lower() == "nan":
                continue
                
            # A. Standard Lookup next to source column (Source field is not "Add new Field")
            if data_type == "lookup" and source_field.lower() != "add new field":
                lookups_attempted += 1
                orig_vals = get_original_col_values(source_field)
                
                if orig_vals is not None:
                    # Clean up rule sheet name
                    sheet_ref = master_sheet_val if master_sheet_val and master_sheet_val.lower() != "nan" else rule_val
                    sheet_name = sheet_ref.split(">")[0].strip().lower() if ">" in sheet_ref else sheet_ref.lower()
                    
                    master_df = master_sheets.get(sheet_name)
                    lookup_map = {}
                    reverse_lookup_map = {}
                    
                    if master_df is not None:
                        # Match search column and copyable column case-insensitively and fuzzy-match
                        search_col = next((c for c in master_df.columns if str(c).strip().lower() == search_col_name.lower()), None)
                        if search_col is None and search_col_name:
                            # Fuzzy matching for typos or slight column variation
                            search_col = next((c for c in master_df.columns if search_col_name.lower() in str(c).lower() or str(c).lower() in search_col_name.lower()), None)
                            
                        # Robust lookup of copy column matching
                        copy_col = None
                        if copyable_col_name:
                            copy_col = next((c for c in master_df.columns if str(c).strip().lower() == copyable_col_name.lower()), None)
                            if copy_col is None:
                                copy_col = next((c for c in master_df.columns if copyable_col_name.lower() in str(c).lower() or str(c).lower() in copyable_col_name.lower()), None)
                        
                        # Absolute fallbacks
                        if search_col is None or search_col not in master_df.columns:
                            search_col = master_df.columns[0]
                        if copy_col is None or copy_col not in master_df.columns:
                            copy_col = next((c for c in master_df.columns if "id" in str(c).lower()), master_df.columns[-1])
                            
                        for _, r in master_df.iterrows():
                            k = str(r[search_col]).strip().lower()
                            v = r[copy_col]
                            if pd.notna(v):
                                lookup_map[k] = v
                                # Support floating matches e.g. 10016901.0 -> 10016901
                                if k.endswith('.0'):
                                    lookup_map[k[:-2]] = v
                                    
                            if pd.notna(v) and pd.notna(r[search_col]):
                                rev_k = str(v).strip().lower()
                                reverse_lookup_map[rev_k] = r[search_col]
                                if rev_k.endswith('.0'):
                                    reverse_lookup_map[rev_k[:-2]] = r[search_col]
                    
                    new_values = []
                    for val in orig_vals:
                        if val is None or pd.isna(val):
                            new_values.append(None)
                            continue
                            
                        val_str = str(val).strip().lower()
                        val_clean = val_str[:-2] if val_str.endswith('.0') else val_str
                        
                        # 1. Try exact match first on lookup_map
                        matched_val = lookup_map.get(val_clean)
                        
                        # 2. Try exact match on reverse_lookup_map
                        if matched_val is None:
                            matched_val = reverse_lookup_map.get(val_clean)
                        
                        # 3. Try fallback substring containment match on lookup_map
                        if matched_val is None:
                            for master_k, master_v in lookup_map.items():
                                if val_clean in master_k or master_k in val_clean:
                                    matched_val = master_v
                                    break
                                    
                        # 4. Try fallback substring containment match on reverse_lookup_map
                        if matched_val is None:
                            for master_k, master_v in reverse_lookup_map.items():
                                if val_clean in master_k or master_k in val_clean:
                                    matched_val = master_v
                                    break
                                    
                        if matched_val is not None:
                            new_values.append(matched_val)
                            lookups_successful += 1
                        else:
                            new_values.append(None)
                            
                    # Find original column index to insert the lookup next to it
                    orig_idx = -1
                    for idx, c in enumerate(columns_list):
                        if c["name"].lower() == source_field.lower() and c["type"] == "original":
                            orig_idx = idx
                            break
                            
                    if orig_idx != -1:
                        # Insert right next to the original column with the identical source header name
                        columns_list.insert(orig_idx + 1, {
                            "name": source_field,
                            "values": new_values,
                            "type": "lookup",
                            "meta": f"Lookup ({target_field})"
                        })
                        
            # B. Relational Lookup with "Add new Field"
            elif data_type == "lookup" and source_field.lower() == "add new field":
                lookups_attempted += 1
                new_fields_added += 1
                
                sheet_ref = master_sheet_val if master_sheet_val and master_sheet_val.lower() != "nan" else rule_val
                sheet_name = sheet_ref.split(">")[0].strip().lower() if ">" in sheet_ref else sheet_ref.lower()
                copyable_col = sheet_ref.split(">")[1].strip() if ">" in sheet_ref else copyable_col_name
                
                # Dynamic Relationship Resolver: dynamically find matching source columns, keys, and values.
                # Completely schema-agnostic and will adapt to any uploaded files dynamically.
                master_df = master_sheets.get(sheet_name)
                new_values = []
                
                if master_df is not None:
                    # 1. Determine target column to copy in master sheet (copy_col)
                    copy_col = None
                    if ">" in sheet_ref:
                        copyable_col_from_ref = sheet_ref.split(">")[1].strip()
                        copy_col = next((c for c in master_df.columns if str(c).strip().lower() == copyable_col_from_ref.lower() or copyable_col_from_ref.lower() in str(c).lower()), None)
                    
                    if copy_col is None and copyable_col_name:
                        copy_col = next((c for c in master_df.columns if str(c).strip().lower() == copyable_col_name.lower() or copyable_col_name.lower() in str(c).lower()), None)
                    
                    if copy_col is None:
                        copy_col = next((c for c in master_df.columns if "id" in str(c).lower()), master_df.columns[-1])
                    
                    # 2. Determine search key column in source data (source_key_col_name) and its values
                    source_key_col_name = None
                    master_pk_col_name = None
                    
                    # Prioritize explicitly defined search column for the current row
                    curr_pk = str(rule.get(col_mapping.get("search_col", ""))).strip()
                    if curr_pk and curr_pk.lower() != "nan":
                        master_pk_col_name = curr_pk

                    # Method A: Try to find another row in mapping logic that maps this same master sheet with a non-"Add new Field" source
                    for _, other_rule in logic_df.iterrows():
                        other_src = str(other_rule.get(col_mapping.get("source_field", ""))).strip()
                        if other_src and other_src.lower() != "nan" and other_src.lower() != "add new field":
                            other_sheet_val, other_rule_val = get_rule_and_sheet_vals(other_rule)
                            other_sheet_ref = other_sheet_val if other_sheet_val and other_sheet_val.lower() != "nan" else other_rule_val
                            other_sheet_name = other_sheet_ref.split(">")[0].strip().lower() if ">" in other_sheet_ref else other_sheet_ref.lower()
                            
                            if other_sheet_name == sheet_name:
                                source_key_col_name = other_src
                                if master_pk_col_name is None:
                                    master_pk_col_name = str(other_rule.get(col_mapping.get("search_col", ""))).strip()
                                break
                    
                    # Method B: Parse the comment field of the current row to find a matching source column name
                    if source_key_col_name is None:
                        comment_val = str(rule.get(col_mapping.get("comments", ""))).strip().lower()
                        if comment_val and comment_val != "nan":
                            sorted_source_cols = sorted([col["name"] for col in columns_list if col["type"] == "original"], key=len, reverse=True)
                            for s_col in sorted_source_cols:
                                if s_col.lower() in comment_val:
                                    source_key_col_name = s_col
                                    break
                                    
                    # Method C: Special defaults based on sheet signatures
                    if source_key_col_name is None and "pricebook" in sheet_name:
                        funding_exists = any(col["name"].strip().lower() == "funding source" for col in columns_list)
                        if funding_exists:
                            source_key_col_name = next(col["name"] for col in columns_list if col["name"].strip().lower() == "funding source")
                    
                    # Resolve values
                    if source_key_col_name is not None:
                        source_key_vals = get_original_col_values(source_key_col_name)
                    else:
                        # Method D: Fallback to schema-agnostic ID resolver
                        source_key_vals = get_carelink_id_values()
                        
                    # 3. Determine search column in the master sheet (master_pk_col)
                    if master_pk_col_name is None or master_pk_col_name.lower() == "nan" or not master_pk_col_name:
                        # In case neither curr_pk nor Method A succeeded, try again with default curr_pk (redundancy fallback)
                        curr_pk = str(rule.get(col_mapping.get("search_col", ""))).strip()
                        if curr_pk and curr_pk.lower() != "nan":
                            master_pk_col_name = curr_pk
                            
                    master_pk_col = None
                    if master_pk_col_name:
                        master_pk_col = next((c for c in master_df.columns if str(c).strip().lower() == master_pk_col_name.lower() or master_pk_col_name.lower() in str(c).lower()), None)
                    
                    if master_pk_col is None and source_key_col_name:
                        master_pk_col = next((c for c in master_df.columns if str(c).strip().lower() == source_key_col_name.lower() or source_key_col_name.lower() in str(c).lower()), None)
                    
                    if master_pk_col is None:
                        if "pricebook" in sheet_name:
                            master_pk_col = next((c for c in master_df.columns if "name" in str(c).lower() or "excel sheet" in str(c).lower() or "sheet name" in str(c).lower()), master_df.columns[-1])
                        elif "accountcarelinkid" in sheet_name:
                            master_pk_col = next((c for c in master_df.columns if "carelink" in str(c).lower()), "Carelink_ID__c")
                        else:
                            master_pk_col = next((c for c in master_df.columns if "id" in str(c).lower() or "key" in str(c).lower()), master_df.columns[0])
                            
                    if master_pk_col is None or master_pk_col not in master_df.columns:
                        master_pk_col = master_df.columns[0]
                        
                    # 4. Build Lookup Map
                    lookup_map = {}
                    reverse_lookup_map = {}
                    for _, r in master_df.iterrows():
                        k = str(r[master_pk_col]).strip().lower()
                        v = r[copy_col]
                        if pd.notna(v):
                            if k.endswith('.0'): k = k[:-2]
                            lookup_map[k] = v
                            
                        if pd.notna(v) and pd.notna(r[master_pk_col]):
                            rev_k = str(v).strip().lower()
                            if rev_k.endswith('.0'): rev_k = rev_k[:-2]
                            reverse_lookup_map[rev_k] = r[master_pk_col]
                        
                    # 5. Perform Mapping
                    if source_key_vals is not None:
                        for val in source_key_vals:
                            if pd.isna(val) or val is None:
                                new_values.append(None)
                            else:
                                val_str = str(val).strip().lower()
                                val_clean = val_str[:-2] if val_str.endswith('.0') else val_str
                                
                                # 1. Try exact match first on lookup_map
                                matched = lookup_map.get(val_clean)
                                
                                # 2. Try exact match on reverse_lookup_map
                                if matched is None:
                                    matched = reverse_lookup_map.get(val_clean)
                                    
                                # 3. Try fallback substring match on lookup_map
                                if matched is None:
                                    for master_k, master_v in lookup_map.items():
                                        if val_clean in master_k or master_k in val_clean:
                                            matched = master_v
                                            break
                                            
                                # 4. Try fallback substring match on reverse_lookup_map
                                if matched is None:
                                    for master_k, master_v in reverse_lookup_map.items():
                                        if val_clean in master_k or master_k in val_clean:
                                            matched = master_v
                                            break
                                            
                                if matched is not None:
                                    new_values.append(matched)
                                    lookups_successful += 1
                                else:
                                    new_values.append(None)
                    else:
                        new_values = [None] * len(columns_list[0]["values"])
                else:
                    new_values = [None] * len(columns_list[0]["values"])
                    
                columns_list.append({
                    "name": target_field,
                    "values": new_values,
                    "type": "new_lookup",
                    "meta": f"Lookup ({sheet_name})"
                })
                
            # C. Date and Date & Time conversions
            elif data_type in ["date", "date&time", "date & time"]:
                found = False
                for c in columns_list:
                    if c["name"].lower() == source_field.lower() and c["type"] == "original":
                        formatted_vals = []
                        for val in c["values"]:
                            fmt_val = format_date_value(val, rule_val)
                            formatted_vals.append(fmt_val)
                            if fmt_val != val:
                                dates_formatted += 1
                        c["values"] = formatted_vals
                        c["name"] = target_field # Update column name to target field
                        found = True
                        break
                        
            # D. Add new Field with constant/default values
            elif source_field.lower() == "add new field" and rule_val and rule_val.lower() != "nan":
                new_fields_added += 1
                num_rows = len(columns_list[0]["values"]) if len(columns_list) > 0 else 0
                
                # Standardize constant strings (e.g. "Default to Nextt" -> "Nextt")
                const_val = rule_val
                if str(rule_val).lower().startswith("default to "):
                    const_val = str(rule_val)[11:].strip()
                    
                new_values = [const_val] * num_rows
                
                columns_list.append({
                    "name": target_field,
                    "values": new_values,
                    "type": "new_constant",
                    "meta": "Constant"
                })
                
        # Resolve duplicate names by prefixing original columns with "__" if a mapped column has the same name
        # Mapped columns are those with type != "original"
        mapped_names = {col["name"].strip().lower(): col["name"] for col in columns_list if col["type"] != "original"}
        
        for col in columns_list:
            if col["type"] == "original":
                col_name_lower = col["name"].strip().lower()
                if col_name_lower in mapped_names:
                    orig_name = col["name"]
                    col["name"] = f"__{orig_name}"
                    print(f"Renaming original column '{orig_name}' to '__{orig_name}' due to collision with mapped column '{mapped_names[col_name_lower]}'.")

        # Resolve target field for each column and prepend it as the first row of data
        source_to_target = {}
        for _, rule in logic_df.iterrows():
            src = get_val(rule, "source_field")
            tgt = get_val(rule, "target_field")
            if src and src.lower() != "nan" and tgt and tgt.lower() != "nan":
                source_to_target[src.lower().strip()] = tgt.strip()
                
        for col in columns_list:
            if col["name"] == "uniqueID":
                target_field_val = "uniqueID"
            elif col["type"] in ["new_lookup", "new_constant"]:
                target_field_val = col["name"]
            else:
                # Resolve using clean original name
                name_clean = col["name"]
                if name_clean.startswith("__"):
                    name_clean = name_clean[2:]
                
                name_clean_lower = name_clean.lower().strip()
                
                # Check stored target_field
                if "target_field" in col:
                    target_field_val = col["target_field"]
                # Check matching source field
                elif name_clean_lower in source_to_target:
                    target_field_val = source_to_target[name_clean_lower]
                # Check standard lookup meta target field
                elif col.get("meta", "").startswith("Lookup (") and col["meta"].endswith(")"):
                    target_field_val = col["meta"][8:-1]
                # Fallback to the clean name itself if it matches any target field in mapping logic
                else:
                    matched_tgt = None
                    for _, rule in logic_df.iterrows():
                        tgt = get_val(rule, "target_field")
                        if tgt.lower().strip() == name_clean_lower:
                            matched_tgt = tgt
                            break
                    target_field_val = matched_tgt if matched_tgt else ""
            
            # Store resolved target field
            col["target_field_val"] = target_field_val
            # Prepend the original Source column name as the first value in the values list
            col["values"].insert(0, col["name"])

        # 6. Save target preview.xlsx
        print(f"Saving finalized preview data to {output_path}...")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        series_list = []
        col_names = []
        for col in columns_list:
            series_list.append(pd.Series(col["values"]))
            col_names.append(col["target_field_val"])
            
        final_df = pd.concat(series_list, axis=1)
        final_df.columns = col_names
        
        # Save beautifully formatted output file using openpyxl
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="Migration Preview")
            
            # Format sheet with openpyxl
            workbook = writer.book
            worksheet = writer.sheets["Migration Preview"]
            
            # Enable gridlines explicitly
            worksheet.views.sheetView[0].showGridLines = True
            
            # Styling definitions
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Fonts
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            cell_font = Font(name="Segoe UI", size=10, color="1E293B")
            id_font = Font(name="Consolas", size=9.5, color="1A73E8") # Royal Blue console text for Salesforce IDs
            constant_font = Font(name="Segoe UI", size=10, italic=True, color="7C3AED") # Custom Purple italic for constant fields
            
            # Fills
            header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Executive dark slate gray
            zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Premium modern light gray/blue zebra shading
            
            # Border styles
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            
            # Format Headers
            for col_idx in range(1, len(columns_list) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                # Overwrite value with clean target field name to strip pandas .1, .2 duplicate column suffixes
                cell.value = columns_list[col_idx - 1]["target_field_val"]
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            worksheet.row_dimensions[1].height = 28
            
            # Format Rows
            for row_idx in range(2, len(final_df) + 2):
                worksheet.row_dimensions[row_idx].height = 20
                is_even = (row_idx % 2 == 0)
                
                is_subheader_row = (row_idx == 2)
                
                for col_idx in range(1, len(columns_list) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    
                    if is_subheader_row:
                        # Style Subheader Row with premium styling (contains original Source column names)
                        cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # premium slate-200 fill
                        cell.font = Font(name="Segoe UI", size=10, bold=True, italic=True, color="1E293B") # dark slate bold italic
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        continue
                        
                    if is_even:
                        cell.fill = zebra_fill
                        
                    col_meta = columns_list[col_idx - 1]
                    col_type = col_meta["type"]
                    
                    val_str = str(cell.value or "")
                    
                    if col_type in ["lookup", "new_lookup"]:
                        cell.font = id_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_type == "new_constant":
                        cell.font = constant_font
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.font = cell_font
                        # Center align short integers, codes or numbers
                        if val_str.isdigit() or len(val_str) <= 10:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            
            # Automatically fit column widths dynamically
            for col in worksheet.columns:
                max_len = 0
                for cell in col:
                    val = str(cell.value or "")
                    if len(val) > max_len:
                        max_len = len(val)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        print("Preview Excel beautifully saved and styled successfully!")
        
        # 7. Extract first 50 rows for JSON preview
        preview_rows = []
        num_rows_to_return = min(len(final_df), 50)
        
        for r_idx in range(num_rows_to_return):
            row_dict = {}
            for col_idx, col in enumerate(columns_list):
                row_dict[f"col_{col_idx}"] = col["values"][r_idx]
            preview_rows.append(row_dict)
            
        ui_columns = []
        for col_idx, col in enumerate(columns_list):
            ui_columns.append({
                "key": f"col_{col_idx}",
                "name": col["target_field_val"],
                "type": col["type"],
                "meta": col.get("meta", "")
            })
            
        return {
            "success": True,
            "columns": ui_columns,
            "rows": preview_rows,
            "summary": {
                "total_rows": len(final_df) - 1, # Exclude target fields row
                "total_columns": len(columns_list),
                "cleaned_columns_count": len(dropped_columns),
                "cleaned_columns_list": dropped_columns,
                "lookups_attempted": lookups_attempted,
                "lookups_successful": lookups_successful,
                "dates_formatted": dates_formatted,
                "new_fields_added": new_fields_added
            }
        }
        
    except Exception as e:
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }


def validate_schema(source_path: str, logic_path: str) -> dict:
    """
    Validates schema between source Excel and mapping logic Excel.

    - Reads all column headers from the first sheet of the source Excel.
    - Reads all values from Column A of the first sheet in the mapping logic file,
      ignoring the header row (Column A header expected to be 'Source fields').
    - Comparison is case-insensitive, ignores spaces and underscores, and trims.
    Returns the structure specified by the frontend.
    """
    try:
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"Source file not found at {source_path}")
        if not os.path.exists(logic_path):
            raise FileNotFoundError(f"Mapping Logic file not found at {logic_path}")

        # Read source headers (first sheet)
        src_df = pd.read_excel(source_path, nrows=0)
        source_columns = [str(c) for c in src_df.columns]

        # Read mapping logic column A (first sheet)
        logic_excel = pd.ExcelFile(logic_path)
        logic_sheet = logic_excel.sheet_names[0]
        # Read only first column (A)
        logic_df = pd.read_excel(logic_path, sheet_name=logic_sheet, usecols=[0])
        # The header row is included as the column name; values are the rest
        mapping_header = list(logic_df.columns)[0]
        mapping_values = logic_df.iloc[:, 0].astype(object).fillna("").tolist()

        # Remove header row if present as first value (some files may include header as first row)
        # But spec says Column A always has header 'Source fields' and ignore header row
        # So drop any value equal to the header (case-insensitive)
        cleaned_mapping = []
        for v in mapping_values:
            vs = str(v).strip()
            if vs == "":
                continue
            if vs.strip().lower() == str(mapping_header).strip().lower():
                continue
            cleaned_mapping.append(vs)

        # Normalization helper
        import re
        def _norm(s: str) -> str:
            if s is None:
                return ""
            s2 = str(s).strip().lower()
            s2 = re.sub(r"[_\s]+", "", s2)
            return s2

        src_norm_map = { _norm(orig): orig for orig in source_columns }
        map_norm_map = { _norm(orig): orig for orig in cleaned_mapping }

        src_norm_set = set(k for k in src_norm_map.keys() if k)
        map_norm_set = set(k for k in map_norm_map.keys() if k)

        matched_norm = src_norm_set & map_norm_set

        missing_norm = map_norm_set - src_norm_set
        additional_norm = src_norm_set - map_norm_set

        missing_fields = [map_norm_map[n] for n in sorted(missing_norm)]
        additional_fields = [src_norm_map[n] for n in sorted(additional_norm)]

        result = {
            "schema_valid": len(missing_fields) == 0,
            "source_field_count": len(source_columns),
            "mapping_field_count": len(cleaned_mapping),
            "matched_field_count": len(matched_norm),
            "missing_fields": missing_fields,
            "additional_fields": additional_fields
        }

        return {"success": True, "result": result}

    except Exception as e:
        traceback.print_exc()
        return {"success": False, "error": str(e), "traceback": traceback.format_exc()}
